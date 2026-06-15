#!/usr/bin/env python3
"""Translate English collocation examples to Traditional Chinese using DeepSeek API."""

import json
import os
import sys
import time
import re
import subprocess
import requests

# ── Configuration ──────────────────────────────────────────────────────
DATA_FILE = "data.json"
EXCEL_FILE = "../01 data_base/collocationData_full.xlsx"
BATCH_SIZE = 20
API_URL = "https://api.deepseek.com/v1/chat/completions"
MODEL = "deepseek-chat"

# ── Get API Key ────────────────────────────────────────────────────────
def get_api_key():
    """Get DeepSeek API key from env var or by sourcing .env file."""
    key = os.environ.get("DEEPSEEK_API_KEY")
    if key and key.strip():
        return key.strip()
    
    # Source .env via bash and print the key
    env_file = os.path.expanduser("~/.hermes/.env")
    if os.path.exists(env_file):
        try:
            result = subprocess.run(
                ["bash", "-c", f"source {env_file} && echo $DEEPSEEK_API_KEY"],
                capture_output=True, text=True, timeout=5
            )
            key = result.stdout.strip()
            if key:
                return key
        except Exception:
            pass
    
    print("ERROR: Cannot find DEEPSEEK_API_KEY")
    sys.exit(1)

API_KEY = get_api_key()
print(f"API key loaded (len={len(API_KEY)}, prefix={API_KEY[:8]}...)")

# ── Load Data ──────────────────────────────────────────────────────────
with open(DATA_FILE, "r", encoding="utf-8") as f:
    data = json.load(f)

print(f"Loaded {len(data)} entries from {DATA_FILE}")

# Find entries needing translation
def needs_example_cn(d):
    return bool((d.get("example") or "").strip()) and not (d.get("example_cn") or "").strip()

def needs_simple_cn(d):
    return bool((d.get("simple") or "").strip()) and not (d.get("simple_cn") or "").strip()

pending_ex = [(i, d) for i, d in enumerate(data) if needs_example_cn(d)]
pending_si = [(i, d) for i, d in enumerate(data) if needs_simple_cn(d)]

print(f"Entries needing example_cn: {len(pending_ex)}")
print(f"Entries needing simple_cn: {len(pending_si)}")

# ── Translation ────────────────────────────────────────────────────────
def translate_batch(texts, field_label="example"):
    """Call DeepSeek API to translate a batch of texts. Returns list of translations."""
    if not texts:
        return []
    
    numbered = "\n".join(f"{j+1}. {t}" for j, t in enumerate(texts))
    prompt = (
        f"請將以下英文例句翻譯成繁體中文，使用台灣用語習慣，翻譯要自然口語。"
        f"只回傳翻譯結果，每行一句，依照編號順序：\n{numbered}"
    )
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_KEY}"
    }
    payload = {
        "model": MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.3,
        "max_tokens": 4096
    }
    
    for attempt in range(3):
        try:
            resp = requests.post(API_URL, headers=headers, json=payload, timeout=120)
            if resp.status_code == 200:
                result = resp.json()
                content = result["choices"][0]["message"]["content"].strip()
                # Parse numbered lines
                lines = []
                for line in content.split("\n"):
                    line = line.strip()
                    if not line:
                        continue
                    # Remove leading number like "1. " or "1." or "1)"
                    line = re.sub(r'^\d+[\.\、\))\s]+', '', line).strip()
                    if line:
                        lines.append(line)
                if len(lines) == len(texts):
                    return lines
                else:
                    print(f"  Got {len(lines)} translations, expected {len(texts)}. Retrying (attempt {attempt+2})...")
                    time.sleep(2)
            elif resp.status_code == 429:
                wait = 10 * (attempt + 1)
                print(f"  Rate limited (429). Waiting {wait}s...")
                time.sleep(wait)
            else:
                print(f"  API error {resp.status_code}: {resp.text[:200]}")
                time.sleep(3)
        except requests.exceptions.Timeout:
            print(f"  Timeout. Retrying (attempt {attempt+2})...")
            time.sleep(5)
        except Exception as e:
            print(f"  Exception: {e}")
            time.sleep(3)
    
    print(f"  FAILED after 3 attempts")
    return None

def save_data():
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def process_batches(pending_items, field_name, label):
    """Process batches, save after each."""
    total = (len(pending_items) + BATCH_SIZE - 1) // BATCH_SIZE
    print(f"\n{'='*60}")
    print(f"Translating {label}: {len(pending_items)} entries in {total} batches")
    print(f"{'='*60}")
    
    failed_batches = []
    
    for batch_num in range(total):
        start = batch_num * BATCH_SIZE
        end = min(start + BATCH_SIZE, len(pending_items))
        batch = pending_items[start:end]
        
        # Skip if all done (resume support)
        all_done = True
        for idx, _ in batch:
            if not (data[idx].get(field_name) or "").strip():
                all_done = False
                break
        
        if all_done:
            continue
        
        texts = [d[field_name.replace("_cn", "")] for _, d in batch]
        print(f"[{batch_num+1}/{total}] Translating {len(texts)} {label}...", end=" ", flush=True)
        
        translations = translate_batch(texts, label)
        if translations:
            for (idx, _), cn in zip(batch, translations):
                data[idx][field_name] = cn
            save_data()
            print(f"OK")
        else:
            failed_batches.append(batch_num)
            print(f"FAILED")
        
        time.sleep(0.3)  # Small delay between batches
    
    if failed_batches:
        print(f"\nWARNING: {len(failed_batches)} batches failed: {failed_batches}")
        return False
    return True

# ── Process ────────────────────────────────────────────────────────────
success_ex = process_batches(pending_ex, "example_cn", "example_cn")
success_si = process_batches(pending_si, "simple_cn", "simple_cn")

# ── Final Save ─────────────────────────────────────────────────────────
save_data()
print(f"\nData saved to {DATA_FILE}")

# ── Update Excel ───────────────────────────────────────────────────────
print(f"\nUpdating Excel: {EXCEL_FILE}")
try:
    import openpyxl
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Find column headers
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip()] = col
    
    print(f"  Found {len(headers)} columns: {list(headers.keys())[:20]}...")
    
    # Determine key columns
    cn1_col = None
    cn2_col = None
    phrase_col = None
    
    for name, col in headers.items():
        if name == "中譯1":
            cn1_col = col
        elif name == "中譯2":
            cn2_col = col
        elif name in ("搭配詞", "phrase", "Phrase"):
            phrase_col = col
    
    if not phrase_col:
        # Try to find phrase column by checking row 2 values against first data entry
        first_phrase = data[0]["phrase"]
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=2, column=col).value
            if val and str(val).strip() == first_phrase:
                phrase_col = col
                break
    
    print(f"  phrase_col={phrase_col}, cn1_col={cn1_col}, cn2_col={cn2_col}")
    
    if not phrase_col:
        print("  WARNING: Cannot find phrase column in Excel")
    else:
        # Build phrase -> data mapping
        data_map = {d["phrase"].strip(): d for d in data}
        
        updated_cn1 = 0
        updated_cn2 = 0
        
        for row in range(2, ws.max_row + 1):
            phrase_val = ws.cell(row=row, column=phrase_col).value
            if phrase_val:
                phrase_str = str(phrase_val).strip()
                d = data_map.get(phrase_str)
                if d:
                    if cn1_col and d.get("example_cn"):
                        ws.cell(row=row, column=cn1_col).value = d["example_cn"]
                        updated_cn1 += 1
                    if cn2_col and d.get("simple_cn"):
                        ws.cell(row=row, column=cn2_col).value = d["simple_cn"]
                        updated_cn2 += 1
        
        print(f"  Updated 中譯1: {updated_cn1} rows")
        print(f"  Updated 中譯2: {updated_cn2} rows")
    
    wb.save(EXCEL_FILE)
    print(f"  Excel saved: {EXCEL_FILE}")
except Exception as e:
    print(f"  Excel error: {e}")
    import traceback
    traceback.print_exc()

# ── Summary ────────────────────────────────────────────────────────────
remaining_ex = sum(1 for d in data if needs_example_cn(d))
remaining_si = sum(1 for d in data if needs_simple_cn(d))
print(f"\n{'='*60}")
print(f"FINAL SUMMARY:")
print(f"  Total entries: {len(data)}")
print(f"  Remaining without example_cn: {remaining_ex}")
print(f"  Remaining without simple_cn: {remaining_si}")
print(f"  Complete!" if remaining_ex == 0 and remaining_si == 0 else f"  Some entries still pending!")
